"""
Builds a downloadable .xlsx file in-memory for any of the 6 report types,
given the already-computed report dict from the relevant service function.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)


def _write_header_row(ws, headers: list[str], row_num: int = 1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def export_revenue_summary(report: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Summary"
    ws.cell(row=1, column=1, value=f"Agilent — DETAILED REVENUE SUMMARY | [{report['month']}]").font = TITLE_FONT

    _write_header_row(ws, ["Particulars", "Amount (₹)", "CGST @ 9%", "SGST @ 9%", "Total"], row_num=3)

    rows = [
        ("Total Trip Amount", report["total_trip_amount"], None, None, None),
        ("(+) Maintenance Charges", report["maintenance_charges"], None, None, None),
        ("(+) Creche Duty Charges", report["creche_duty_charges"], None, None, None),
        ("(+) Odd Hours Cab Cost", report["odd_hours_cab_cost"], None, None, None),
        ("Grand Total (Billable Trip Charges)", report["grand_total_billable"], None, None, None),
        ("(-) Amount Recovered from Employees", report["amount_recovered_from_employees"], None, None, report["amount_recovered_from_employees"]),
        ("Taxable Trip Amount", report["taxable_trip_amount"], report["taxable_trip_cgst"], report["taxable_trip_sgst"], report["taxable_trip_total"]),
        ("(+) Manpower Charges", report["manpower_charges"], report["manpower_cgst"], report["manpower_sgst"], report["manpower_total"]),
        ("(+) Technology Cost Recovery", report["technology_cost_recovery"], report["technology_cgst"], report["technology_sgst"], report["technology_total"]),
        ("(+) Dashcam Subscription Recovery", report["dashcam_subscription_recovery"], report["dashcam_cgst"], report["dashcam_sgst"], report["dashcam_total"]),
        ("(+) Razorpay Transaction Fee Recovery", report["razorpay_fee_recovery"], report["razorpay_cgst"], report["razorpay_sgst"], report["razorpay_total"]),
        ("(+) Spot Rental Revenue", report["spot_rental_revenue"], report["spot_rental_cgst"], report["spot_rental_sgst"], report["spot_rental_total"]),
        ("Total Taxable Amount", report["total_taxable_amount"], report["cgst_total"], report["sgst_total"],
         report["net_amount_payable_by_agilent"]),
        ("Net Amount Payable by Agilent", report["net_amount_payable_by_agilent"], None, None, None),
        ("TOTAL REVENUE", report["total_revenue"], None, None, None),
    ]

    for r_idx, (particulars, amount, cgst, sgst, total) in enumerate(rows, start=4):
        ws.cell(row=r_idx, column=1, value=particulars)
        ws.cell(row=r_idx, column=2, value=round(amount, 2))
        ws.cell(row=r_idx, column=3, value=round(cgst, 2) if cgst is not None else "-")
        ws.cell(row=r_idx, column=4, value=round(sgst, 2) if sgst is not None else "-")
        ws.cell(row=r_idx, column=5, value=round(total, 2) if total is not None else "-")

    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 38 if col == "A" else 18

    return _save_to_buffer(wb)


def export_revenue_mix(report: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Mix"
    ws.cell(row=1, column=1, value=f"Agilent — REVENUE MIX BY SOURCE | [{report['month']}]").font = TITLE_FONT

    _write_header_row(ws, ["Revenue Source", "Amount (₹)", "% of Total Revenue"], row_num=3)

    for r_idx, item in enumerate(report["items"], start=4):
        ws.cell(row=r_idx, column=1, value=item["revenue_source"])
        ws.cell(row=r_idx, column=2, value=round(item["amount"], 2))
        ws.cell(row=r_idx, column=3, value=f"{item['percent_of_total']:.2f}%")

    last_row = 4 + len(report["items"])
    ws.cell(row=last_row, column=1, value="Total Revenue").font = Font(bold=True)
    ws.cell(row=last_row, column=2, value=round(report["total_revenue"], 2)).font = Font(bold=True)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    return _save_to_buffer(wb)


def export_vehicle_breakup(report: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle-wise Breakup"
    ws.cell(row=1, column=1, value=f"Agilent — VEHICLE-WISE REVENUE BREAKUP | [{report['month']}]").font = TITLE_FONT

    headers = ["VehicleNumber", "Ownership", "TripDataAmount", "SpotRentel",
               "MaintainenceVehAmount", "ChildCabAmount", "BackupCabsAmount", "GrandTotal", "% of Total"]
    _write_header_row(ws, headers, row_num=3)

    for r_idx, row in enumerate(report["rows"], start=4):
        ws.cell(row=r_idx, column=1, value=row["vehicle_number"])
        ws.cell(row=r_idx, column=2, value=row["ownership"])
        ws.cell(row=r_idx, column=3, value=round(row["trip_data_amount"], 2))
        ws.cell(row=r_idx, column=4, value=round(row["spot_rental"], 2))
        ws.cell(row=r_idx, column=5, value=round(row["maintenance_veh_amount"], 2))
        ws.cell(row=r_idx, column=6, value=round(row["child_cab_amount"], 2))
        ws.cell(row=r_idx, column=7, value=round(row["backup_cabs_amount"], 2))
        ws.cell(row=r_idx, column=8, value=round(row["grand_total"], 2))
        ws.cell(row=r_idx, column=9, value=f"{row['percent_of_total']:.2f}%")

    last_row = 4 + len(report["rows"])
    ws.cell(row=last_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=last_row, column=3, value=round(report["grand_total_trip_data"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=round(report["grand_total_spot_rental"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=round(report["grand_total_maintenance"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=6, value=round(report["grand_total_child_cab"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=7, value=round(report["grand_total_backup_cab"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=8, value=round(report["grand_total_overall"], 2)).font = Font(bold=True)

    for col, width in zip("ABCDEFGHI", [14, 12, 16, 14, 18, 14, 16, 14, 12]):
        ws.column_dimensions[col].width = width

    return _save_to_buffer(wb)


def export_ownership_breakup(report: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ownership Breakup"
    ws.cell(row=1, column=1, value=f"Agilent — REVENUE BY OWNERSHIP TYPE | [{report['month']}]").font = TITLE_FONT

    _write_header_row(ws, ["Ownership Type", "Total Revenue (₹)", "% of Total"], row_num=3)

    for r_idx, row in enumerate(report["rows"], start=4):
        ws.cell(row=r_idx, column=1, value=row["ownership_type"])
        ws.cell(row=r_idx, column=2, value=round(row["total_revenue"], 2))
        ws.cell(row=r_idx, column=3, value=f"{row['percent_of_total']:.2f}%")

    last_row = 4 + len(report["rows"])
    ws.cell(row=last_row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=last_row, column=2, value=round(report["total"], 2)).font = Font(bold=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14

    return _save_to_buffer(wb)


def export_vehicle_revenue_summary(report: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Revenue Summary"
    ws.cell(row=1, column=1, value=f"Agilent — VEHICLE REVENUE SUMMARY | [{report['month']}]").font = TITLE_FONT

    headers = ["VehicleNumber", "Ownership", "TripDataAmount", "SpotRentel",
               "MaintainenceVehAmount", "ChildCabAmount", "BackupCabsAmount", "GrandTotal"]
    _write_header_row(ws, headers, row_num=3)

    for r_idx, row in enumerate(report["rows"], start=4):
        ws.cell(row=r_idx, column=1, value=row["vehicle_number"])
        ws.cell(row=r_idx, column=2, value=row["ownership"])
        ws.cell(row=r_idx, column=3, value=round(row["trip_data_amount"], 2))
        ws.cell(row=r_idx, column=4, value=round(row["spot_rental"], 2))
        ws.cell(row=r_idx, column=5, value=round(row["maintenance_veh_amount"], 2))
        ws.cell(row=r_idx, column=6, value=round(row["child_cab_amount"], 2))
        ws.cell(row=r_idx, column=7, value=round(row["backup_cabs_amount"], 2))
        ws.cell(row=r_idx, column=8, value=round(row["grand_total"], 2))

    last_row = 4 + len(report["rows"])
    ws.cell(row=last_row, column=1, value="Grand Total").font = Font(bold=True)
    ws.cell(row=last_row, column=3, value=round(report["grand_total_trip_data"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=round(report["grand_total_spot_rental"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=round(report["grand_total_maintenance"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=6, value=round(report["grand_total_child_cab"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=7, value=round(report["grand_total_backup_cab"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=8, value=round(report["grand_total_overall"], 2)).font = Font(bold=True)

    for col, width in zip("ABCDEFGH", [14, 12, 16, 14, 18, 14, 16, 14]):
        ws.column_dimensions[col].width = width

    return _save_to_buffer(wb)


def export_pnl_summary(report: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "PNL Summary"
    ws.cell(row=1, column=1, value=f"Agilent — MIS REPORT | [{report['month']}]").font = TITLE_FONT

    _write_header_row(ws, ["Particulars", "Amount (₹)", "% of Sale", "Notes"], row_num=3)

    r_idx = 4
    ws.cell(row=r_idx, column=1, value="REVENUE").font = Font(bold=True)
    r_idx += 1
    ws.cell(row=r_idx, column=1, value="TOTAL REVENUE")
    ws.cell(row=r_idx, column=2, value=round(report["total_revenue"], 2))
    r_idx += 2

    ws.cell(row=r_idx, column=1, value="EXPENSES").font = Font(bold=True)
    r_idx += 1
    for item in report["expenses"]:
        ws.cell(row=r_idx, column=1, value=item["particulars"])
        ws.cell(row=r_idx, column=2, value=round(item["amount"], 2))
        ws.cell(row=r_idx, column=3, value=f"{item['percent_of_sale']:.2f}%")
        r_idx += 1

    ws.cell(row=r_idx, column=1, value="TOTAL EXPENSES").font = Font(bold=True)
    ws.cell(row=r_idx, column=2, value=round(report["total_expenses"], 2)).font = Font(bold=True)
    r_idx += 2

    ws.cell(row=r_idx, column=1, value="PROFITABILITY").font = Font(bold=True)
    r_idx += 1
    ws.cell(row=r_idx, column=1, value="NET PROFIT / (LOSS)")
    ws.cell(row=r_idx, column=2, value=round(report["net_profit_loss"], 2))
    ws.cell(row=r_idx, column=4, value="Profit / (Loss) Margin")
    ws.cell(row=r_idx, column=3, value=f"{report['net_margin_percent']:.2f}%")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 26

    return _save_to_buffer(wb)


def _save_to_buffer(wb: Workbook) -> BytesIO:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
