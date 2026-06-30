"""
Builds a downloadable .xlsx file in-memory for each of the 3 Air India
report types, given the already-computed report dict.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)


def _write_header_row(ws, headers: list[str], row_num: int = 1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def export_revenue_summary(report: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Summary"
    ws.cell(row=1, column=1, value=f"Agilent (Air India) — DETAILED REVENUE SUMMARY | [{report['month']}]").font = TITLE_FONT

    _write_header_row(ws, ["Particulars", "Amount (₹)"], row_num=3)

    rows = [
        ("Trip_Amount_TERMINAL-3", report["trip_amount_t3"]),
        ("TollAmount_TERMINAL-3", report["toll_amount_t3"]),
        ("MCD", report["mcd"]),
        ("Terminal-3 Driver Penalty", -report["t3_driver_penalty"]),
        ("Terminal-3 Employee Penalty", -report["t3_employee_penalty"]),
        ("Total Of Terminal3", report["total_of_terminal3"]),
        ("Trip_Amount_AIAA", report["trip_amount_aiaa"]),
        ("TollAmount_AIAA", report["toll_amount_aiaa"]),
        ("AIAA Driver Penalty", -report["aiaa_driver_penalty"]),
        ("AIAA Employee Penalty", -report["aiaa_employee_penalty"]),
        ("Total of AIAA", report["total_of_aiaa"]),
        ("Total Taxable Amount", report["total_taxable_amount"]),
        ("GST (5% on T-3 + 5% on AIAA)", report["gst_total"]),
        ("TOTAL REVENUE (Net Amount Payable by Agilent)", report["total_revenue"]),
    ]

    for r_idx, (particulars, amount) in enumerate(rows, start=4):
        ws.cell(row=r_idx, column=1, value=particulars)
        ws.cell(row=r_idx, column=2, value=round(amount, 2))

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18

    return _save_to_buffer(wb)


def export_vehicle_revenue_summary(report: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Revenue Summary"
    ws.cell(row=1, column=1, value=f"Agilent (Air India) — VEHICLE REVENUE SUMMARY | [{report['month']}]").font = TITLE_FONT

    headers = ["VehicleNumber", "Ownership", "Vehtype", "T-3Amount", "AIAAamount", "Total"]
    _write_header_row(ws, headers, row_num=3)

    for r_idx, row in enumerate(report["rows"], start=4):
        ws.cell(row=r_idx, column=1, value=row["vehicle_number"])
        ws.cell(row=r_idx, column=2, value=row["ownership"])
        ws.cell(row=r_idx, column=3, value=row["veh_type"])
        ws.cell(row=r_idx, column=4, value=round(row["t3_amount"], 2))
        ws.cell(row=r_idx, column=5, value=round(row["aiaa_amount"], 2))
        ws.cell(row=r_idx, column=6, value=round(row["total"], 2))

    last_row = 4 + len(report["rows"])
    ws.cell(row=last_row, column=1, value="Grand Total").font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=round(report["grand_total_t3"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=round(report["grand_total_aiaa"], 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=6, value=round(report["grand_total_overall"], 2)).font = Font(bold=True)

    for col, width in zip("ABCDEF", [16, 12, 14, 16, 16, 16]):
        ws.column_dimensions[col].width = width

    return _save_to_buffer(wb)


def export_pnl_summary(report: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "PNL Summary"
    ws.cell(row=1, column=1, value=f"Agilent (Air India) — MIS REPORT | [{report['month']}]").font = TITLE_FONT

    _write_header_row(ws, ["Particulars", "Amount (₹)", "% of Sale", "Notes"], row_num=3)

    r_idx = 4
    ws.cell(row=r_idx, column=1, value="REVENUE").font = Font(bold=True)
    r_idx += 1
    ws.cell(row=r_idx, column=1, value="TOTAL REVENUE")
    ws.cell(row=r_idx, column=2, value=round(report["total_revenue"], 2))
    r_idx += 2

    ws.cell(row=r_idx, column=1, value="EXPENSES").font = Font(bold=True)
    r_idx += 1

    ws.cell(row=r_idx, column=1, value="Vehicle Operating Costs (Own Vehicles)").font = Font(italic=True)
    r_idx += 1
    for item in [report["fuel"], report["vehicle_maintenance_cost"], report["drivers_salaries"], report["vehicle_emi"]]:
        ws.cell(row=r_idx, column=1, value=item["particulars"])
        ws.cell(row=r_idx, column=2, value=round(item["amount"], 2))
        ws.cell(row=r_idx, column=3, value=f"{item['percent_of_sale']:.2f}%")
        r_idx += 1

    ws.cell(row=r_idx, column=1, value="Technology & Software").font = Font(italic=True)
    r_idx += 1
    item = report["razorpay_transaction_fee"]
    ws.cell(row=r_idx, column=1, value=item["particulars"])
    ws.cell(row=r_idx, column=2, value=round(item["amount"], 2))
    ws.cell(row=r_idx, column=3, value=f"{item['percent_of_sale']:.2f}%")
    r_idx += 1

    ws.cell(row=r_idx, column=1, value="Operational Expenses").font = Font(italic=True)
    r_idx += 1
    for item in [report["mcd_state_taxes_toll_parking"], report["vendor_payment"], report["employee_salary"]]:
        ws.cell(row=r_idx, column=1, value=item["particulars"])
        ws.cell(row=r_idx, column=2, value=round(item["amount"], 2))
        ws.cell(row=r_idx, column=3, value=f"{item['percent_of_sale']:.2f}%")
        r_idx += 1

    ws.cell(row=r_idx, column=1, value="Taxes").font = Font(italic=True)
    r_idx += 1
    item = report["gst"]
    ws.cell(row=r_idx, column=1, value=item["particulars"])
    ws.cell(row=r_idx, column=2, value=round(item["amount"], 2))
    ws.cell(row=r_idx, column=3, value=f"{item['percent_of_sale']:.2f}%")
    r_idx += 1

    ws.cell(row=r_idx, column=1, value="TOTAL EXPENSES").font = Font(bold=True)
    ws.cell(row=r_idx, column=2, value=round(report["total_expenses"], 2)).font = Font(bold=True)
    r_idx += 2

    ws.cell(row=r_idx, column=1, value="PROFITABILITY").font = Font(bold=True)
    r_idx += 1
    ws.cell(row=r_idx, column=1, value="NET PROFIT / (LOSS)")
    ws.cell(row=r_idx, column=2, value=round(report["net_profit_loss"], 2))
    ws.cell(row=r_idx, column=4, value="Profit / (Loss) Margin")
    ws.cell(row=r_idx, column=3, value=f"{report['net_margin_percent']:.2f}%")

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 26

    return _save_to_buffer(wb)


def _save_to_buffer(wb: Workbook) -> BytesIO:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
